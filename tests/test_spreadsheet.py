from openpyxl import Workbook, load_workbook

from price_generator.constants import (
    COMBO_COST_PRICE_COLUMN_NAME,
    COMBO_PRODUCT_CODE_COLUMN_NAME,
    COMBO_WEIGHT_COLUMN_NAME,
    COST_PRICE_COLUMN_NAME,
    PRICE_COLUMN_NAME,
    PRODUCT_CODE_COLUMN_NAME,
    SPEC_CODE_COLUMN_NAME,
    WEIGHT_COLUMN_NAME,
)
from price_generator.spreadsheet import (
    append_cached_prices_xlsx,
    import_prices_to_cache_xlsx,
)


def save_workbook(path, header, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def read_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()


def test_import_single_product_prices_to_cache(tmp_path):
    input_path = tmp_path / "single.xlsx"
    cache = {}
    save_workbook(
        input_path,
        [PRODUCT_CODE_COLUMN_NAME, WEIGHT_COLUMN_NAME, COST_PRICE_COLUMN_NAME],
        [["abc", 0.24, 1]],
    )

    summary = import_prices_to_cache_xlsx(input_path, price_cache=cache)

    assert summary == [("Sheet1", 1, 0, 0)]
    assert cache == {"ABC": 2.64}


def test_import_combo_product_prices_to_cache(tmp_path):
    input_path = tmp_path / "combo.xlsx"
    cache = {}
    save_workbook(
        input_path,
        [COMBO_PRODUCT_CODE_COLUMN_NAME, COMBO_WEIGHT_COLUMN_NAME, COMBO_COST_PRICE_COLUMN_NAME],
        [["combo1", 4.9, 5]],
    )

    summary = import_prices_to_cache_xlsx(
        input_path,
        price_cache=cache,
        product_code_column=COMBO_PRODUCT_CODE_COLUMN_NAME,
        weight_column=COMBO_WEIGHT_COLUMN_NAME,
        cost_price_column=COMBO_COST_PRICE_COLUMN_NAME,
    )

    assert summary == [("Sheet1", 1, 0, 0)]
    assert cache == {"COMBO1": 21.26}


def test_append_cached_prices_uses_spec_code_lookup(tmp_path):
    input_path = tmp_path / "new.xlsx"
    output_path = tmp_path / "new_out.xlsx"
    save_workbook(input_path, [SPEC_CODE_COLUMN_NAME], [["abc$$suffix"], ["missing"]])

    summary = append_cached_prices_xlsx(input_path, output_path, price_cache={"ABC": 12.3})
    rows = read_rows(output_path)

    assert summary == [("Sheet1", 1, 1)]
    assert rows[0] == (SPEC_CODE_COLUMN_NAME, PRICE_COLUMN_NAME)
    assert rows[1] == ("abc$$suffix", 12.3)
    assert rows[2][0] == "missing"
