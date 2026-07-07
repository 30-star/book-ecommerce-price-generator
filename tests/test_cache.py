import pytest
from openpyxl import load_workbook

from price_generator.cache import export_price_cache_xlsx, load_price_cache, save_price_cache


def test_price_cache_save_load_normalizes_keys(tmp_path):
    cache_path = tmp_path / "price_cache.json"
    save_price_cache(
        {
            "abc": {"cost_price": 1, "shipping_fee": 1.3, "price": 2.64},
            "": {"cost_price": 99, "shipping_fee": 99, "price": 99},
        },
        cache_path,
    )

    assert load_price_cache(cache_path) == {"ABC": {"cost_price": 1, "shipping_fee": 1.3, "price": 2.64}}


def test_price_cache_rejects_missing_shipping_fee(tmp_path):
    cache_path = tmp_path / "price_cache.json"
    cache_path.write_text('{"prices": {"ABC": {"cost_price": 1, "price": 2.64}}}', encoding="utf-8")

    with pytest.raises(ValueError, match="shipping_fee"):
        load_price_cache(cache_path)


def test_price_cache_export_xlsx(tmp_path):
    output_path = tmp_path / "cache.xlsx"
    count = export_price_cache_xlsx(
        output_path,
        {"ABC": {"cost_price": 1, "shipping_fee": 1.3, "price": 2.64}},
    )

    assert count == 1
    workbook = load_workbook(output_path, read_only=True)
    try:
        rows = list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()

    assert rows == [("商品编码", "成本价", "快递费", "售价"), ("ABC", 1, 1.3, 2.64)]
