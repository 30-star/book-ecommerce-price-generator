import csv
import os
import warnings

from openpyxl import Workbook, load_workbook

from .constants import *
from .headers import build_header_map, find_header_index, get_row_column_value
from .pricing import calculate_price, calculate_shipping_fee, normalize_price_cache_key, normalize_spec_code_for_price_lookup, resolve_price
from .rules import get_row_value, parse_number, should_delete_row

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    category=UserWarning,
    module="openpyxl.styles.stylesheet",
)


def insert_value(row, index, value):
    next_row = list(row)
    while len(next_row) < index:
        next_row.append(None)
    next_row.insert(index, value)
    return next_row


def insert_values(row, index, values):
    next_row = list(row)
    while len(next_row) < index:
        next_row.append(None)

    for offset, value in enumerate(values):
        next_row.insert(index + offset, value)

    return next_row


def report_progress(progress_callback, text, current=None, total=None):
    if not progress_callback:
        return

    percent = None
    if current is not None and total:
        percent = max(0, min(100, int(current * 100 / total)))

    progress_callback(text, percent)


def estimate_xlsx_total_rows(workbook):
    total = 0

    for sheet in workbook.worksheets:
        max_row = sheet.max_row or 0
        if max_row > 1:
            total += max_row - 1

    return total or None


def process_xlsx(
    input_path,
    output_path,
    insert_side="after",
    progress_callback=None,
    shipping_fees=None,
    price_rules=None,
    delete_rules=None,
    price_match_column=DEFAULT_PRICE_MATCH_COLUMN_NAME,
    price_cache=None,
):
    source = load_workbook(input_path, read_only=True, data_only=True)
    target = Workbook(write_only=True)
    price_cache = price_cache if price_cache is not None else {}

    if target.worksheets:
        target.remove(target.worksheets[0])

    summary = []

    try:
        total_rows = estimate_xlsx_total_rows(source)
        processed_rows = 0

        for sheet_index, source_sheet in enumerate(source.worksheets, start=1):
            target_sheet = target.create_sheet(title=source_sheet.title)
            rows = source_sheet.iter_rows(values_only=True)
            header = next(rows, None)

            if header is None:
                target_sheet.append([])
                summary.append((source_sheet.title, 0, 0, 0, 0))
                continue

            headers = list(header)
            header_map = build_header_map(headers)
            weight_index = find_header_index(headers, WEIGHT_COLUMN_NAME)
            product_index = find_header_index(headers, PRODUCT_COLUMN_NAME)
            cost_price_index = find_header_index(headers, COST_PRICE_COLUMN_NAME)
            available_index = find_header_index(headers, AVAILABLE_COLUMN_NAME)
            stock_index = find_header_index(headers, STOCK_COLUMN_NAME)
            product_tag_index = find_header_index(headers, PRODUCT_TAG_COLUMN_NAME)
            supplier_index = find_header_index(headers, SUPPLIER_COLUMN_NAME)
            sales_15_days_index = find_header_index(headers, SALES_15_DAYS_COLUMN_NAME)

            if weight_index == -1:
                raise ValueError(f"工作表“{source_sheet.title}”没有找到“{WEIGHT_COLUMN_NAME}”列。")
            if find_header_index(headers, price_match_column) == -1:
                raise ValueError(f"工作表“{source_sheet.title}”没有找到售价匹配列“{price_match_column}”。")

            insert_index = weight_index if insert_side == "before" else weight_index + 1
            target_sheet.append(insert_values(headers, insert_index, [INSERTED_COLUMN_NAME, PRICE_COLUMN_NAME]))

            kept_rows = 0
            deleted_rows = 0
            matched_prices = 0
            saved_prices = 0

            for row_number, row in enumerate(rows, start=2):
                row_values = list(row)
                processed_rows += 1

                if should_delete_row(row_values, header_map, delete_rules or DEFAULT_DELETE_RULES):
                    deleted_rows += 1
                    if row_number % 500 == 0:
                        report_progress(
                            progress_callback,
                            f"正在处理：{source_sheet.title} 第 {row_number} 行",
                            processed_rows,
                            total_rows,
                        )
                    continue

                weight_value = get_row_value(row_values, weight_index)
                shipping_fee = calculate_shipping_fee(weight_value, shipping_fees)
                price, matched, saved = resolve_price(
                    row_values,
                    header_map,
                    cost_price_index,
                    shipping_fee,
                    price_rules,
                    price_match_column,
                    price_cache,
                )
                if matched:
                    matched_prices += 1
                if saved:
                    saved_prices += 1
                target_sheet.append(insert_values(row_values, insert_index, [shipping_fee, price]))
                kept_rows += 1

                if row_number % 500 == 0:
                    report_progress(
                        progress_callback,
                        f"正在处理：{source_sheet.title} 第 {row_number} 行",
                        processed_rows,
                        total_rows,
                    )

            summary.append((source_sheet.title, kept_rows, deleted_rows, matched_prices, saved_prices))

            report_progress(
                progress_callback,
                f"已处理 {sheet_index}/{len(source.worksheets)} 个工作表",
                processed_rows,
                total_rows,
            )

        report_progress(progress_callback, "正在保存导出文件...", total_rows, total_rows)
        target.save(output_path)
    finally:
        source.close()
        target.close()

    return summary


def process_csv(
    input_path,
    output_path,
    insert_side="after",
    progress_callback=None,
    shipping_fees=None,
    price_rules=None,
    delete_rules=None,
    price_match_column=DEFAULT_PRICE_MATCH_COLUMN_NAME,
    price_cache=None,
):
    price_cache = price_cache if price_cache is not None else {}

    with open(input_path, "r", newline="", encoding="utf-8-sig") as count_file:
        total_rows = max(sum(1 for _ in count_file) - 1, 0)

    with open(input_path, "r", newline="", encoding="utf-8-sig") as source_file:
        sample = source_file.read(4096)
        source_file.seek(0)
        dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
        reader = csv.reader(source_file, dialect)

        header = next(reader, None)
        if header is None:
            raise ValueError("CSV 文件为空。")

        header_map = build_header_map(header)
        weight_index = find_header_index(header, WEIGHT_COLUMN_NAME)
        product_index = find_header_index(header, PRODUCT_COLUMN_NAME)
        cost_price_index = find_header_index(header, COST_PRICE_COLUMN_NAME)
        available_index = find_header_index(header, AVAILABLE_COLUMN_NAME)
        stock_index = find_header_index(header, STOCK_COLUMN_NAME)
        product_tag_index = find_header_index(header, PRODUCT_TAG_COLUMN_NAME)
        supplier_index = find_header_index(header, SUPPLIER_COLUMN_NAME)
        sales_15_days_index = find_header_index(header, SALES_15_DAYS_COLUMN_NAME)

        if weight_index == -1:
            raise ValueError(f"没有找到“{WEIGHT_COLUMN_NAME}”列。")
        if find_header_index(header, price_match_column) == -1:
            raise ValueError(f"没有找到售价匹配列“{price_match_column}”。")

        insert_index = weight_index if insert_side == "before" else weight_index + 1

        with open(output_path, "w", newline="", encoding="utf-8-sig") as target_file:
            writer = csv.writer(target_file)
            writer.writerow(insert_values(header, insert_index, [INSERTED_COLUMN_NAME, PRICE_COLUMN_NAME]))

            kept_rows = 0
            deleted_rows = 0
            matched_prices = 0
            saved_prices = 0
            processed_rows = 0

            for row_number, row in enumerate(reader, start=2):
                processed_rows += 1

                if should_delete_row(row, header_map, delete_rules or DEFAULT_DELETE_RULES):
                    deleted_rows += 1
                    if row_number % 500 == 0:
                        report_progress(progress_callback, f"正在处理 CSV 第 {row_number} 行", processed_rows, total_rows)
                    continue

                weight_value = get_row_value(row, weight_index)
                shipping_fee = calculate_shipping_fee(weight_value, shipping_fees)
                price, matched, saved = resolve_price(
                    row,
                    header_map,
                    cost_price_index,
                    shipping_fee,
                    price_rules,
                    price_match_column,
                    price_cache,
                )
                if matched:
                    matched_prices += 1
                if saved:
                    saved_prices += 1
                writer.writerow(insert_values(row, insert_index, [shipping_fee, price]))
                kept_rows += 1

                if row_number % 500 == 0:
                    report_progress(progress_callback, f"正在处理 CSV 第 {row_number} 行", processed_rows, total_rows)

    report_progress(progress_callback, "CSV 处理完成。", total_rows, total_rows)

    return [("CSV", kept_rows, deleted_rows, matched_prices, saved_prices)]


def append_cached_prices_xlsx(input_path, output_path, progress_callback=None, price_cache=None):
    source = load_workbook(input_path, read_only=True, data_only=True)
    target = Workbook(write_only=True)
    price_cache = price_cache if price_cache is not None else {}

    if target.worksheets:
        target.remove(target.worksheets[0])

    summary = []

    try:
        total_rows = estimate_xlsx_total_rows(source)
        processed_rows = 0

        for sheet_index, source_sheet in enumerate(source.worksheets, start=1):
            target_sheet = target.create_sheet(title=source_sheet.title)
            rows = source_sheet.iter_rows(values_only=True)
            header = next(rows, None)

            if header is None:
                target_sheet.append([PRICE_COLUMN_NAME])
                summary.append((source_sheet.title, 0, 0))
                continue

            headers = list(header)
            header_map = build_header_map(headers)

            if find_header_index(headers, SPEC_CODE_COLUMN_NAME) == -1:
                raise ValueError(f"工作表“{source_sheet.title}”没有找到“{SPEC_CODE_COLUMN_NAME}”列。")

            target_sheet.append(headers + [PRICE_COLUMN_NAME])

            matched_rows = 0
            unmatched_rows = 0

            for row_number, row in enumerate(rows, start=2):
                row_values = list(row)
                processed_rows += 1
                product_code = normalize_spec_code_for_price_lookup(
                    get_row_column_value(row_values, header_map, SPEC_CODE_COLUMN_NAME)
                )
                cached_record = price_cache.get(product_code) if product_code else None
                price = cached_record.get("price") if cached_record else None

                if price is None:
                    unmatched_rows += 1
                else:
                    matched_rows += 1

                target_sheet.append(row_values + [price])

                if row_number % 500 == 0:
                    report_progress(
                        progress_callback,
                        f"正在匹配售价：{source_sheet.title} 第 {row_number} 行",
                        processed_rows,
                        total_rows,
                    )

            summary.append((source_sheet.title, matched_rows, unmatched_rows))
            report_progress(
                progress_callback,
                f"已匹配 {sheet_index}/{len(source.worksheets)} 个工作表",
                processed_rows,
                total_rows,
            )

        report_progress(progress_callback, "正在保存匹配结果...", total_rows, total_rows)
        target.save(output_path)
    finally:
        source.close()
        target.close()

    return summary


def append_cached_prices_csv(input_path, output_path, progress_callback=None, price_cache=None):
    price_cache = price_cache if price_cache is not None else {}

    with open(input_path, "r", newline="", encoding="utf-8-sig") as count_file:
        total_rows = max(sum(1 for _ in count_file) - 1, 0)

    with open(input_path, "r", newline="", encoding="utf-8-sig") as source_file:
        sample = source_file.read(4096)
        source_file.seek(0)
        dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
        reader = csv.reader(source_file, dialect)

        header = next(reader, None)
        if header is None:
            raise ValueError("CSV 文件为空。")

        header_map = build_header_map(header)
        if find_header_index(header, SPEC_CODE_COLUMN_NAME) == -1:
            raise ValueError(f"没有找到“{SPEC_CODE_COLUMN_NAME}”列。")

        with open(output_path, "w", newline="", encoding="utf-8-sig") as target_file:
            writer = csv.writer(target_file)
            writer.writerow(list(header) + [PRICE_COLUMN_NAME])

            matched_rows = 0
            unmatched_rows = 0
            processed_rows = 0

            for row_number, row in enumerate(reader, start=2):
                processed_rows += 1
                product_code = normalize_spec_code_for_price_lookup(
                    get_row_column_value(row, header_map, SPEC_CODE_COLUMN_NAME)
                )
                cached_record = price_cache.get(product_code) if product_code else None
                price = cached_record.get("price") if cached_record else None

                if price is None:
                    unmatched_rows += 1
                else:
                    matched_rows += 1

                writer.writerow(list(row) + [price])

                if row_number % 500 == 0:
                    report_progress(progress_callback, f"正在匹配 CSV 第 {row_number} 行", processed_rows, total_rows)

    report_progress(progress_callback, "CSV 售价匹配完成。", total_rows, total_rows)

    return [("CSV", matched_rows, unmatched_rows)]


def import_prices_to_cache_xlsx(
    input_path,
    progress_callback=None,
    shipping_fees=None,
    price_rules=None,
    price_cache=None,
    product_code_column=PRODUCT_CODE_COLUMN_NAME,
    weight_column=WEIGHT_COLUMN_NAME,
    cost_price_column=COST_PRICE_COLUMN_NAME,
    progress_label="价格库",
):
    source = load_workbook(input_path, read_only=True, data_only=True)
    price_cache = price_cache if price_cache is not None else {}
    summary = []

    try:
        total_rows = estimate_xlsx_total_rows(source)
        processed_rows = 0

        for sheet_index, source_sheet in enumerate(source.worksheets, start=1):
            rows = source_sheet.iter_rows(values_only=True)
            header = next(rows, None)

            if header is None:
                summary.append((source_sheet.title, 0, 0, 0))
                continue

            headers = list(header)
            header_map = build_header_map(headers)
            product_code_index = find_header_index(headers, product_code_column)
            weight_index = find_header_index(headers, weight_column)
            cost_price_index = find_header_index(headers, cost_price_column)

            if product_code_index == -1:
                raise ValueError(f"工作表“{source_sheet.title}”没有找到“{product_code_column}”列。")
            if weight_index == -1:
                raise ValueError(f"工作表“{source_sheet.title}”没有找到“{weight_column}”列。")
            if cost_price_index == -1:
                raise ValueError(f"工作表“{source_sheet.title}”没有找到“{cost_price_column}”列。")

            added_rows = 0
            updated_rows = 0
            skipped_rows = 0

            for row_number, row in enumerate(rows, start=2):
                row_values = list(row)
                processed_rows += 1
                product_code = normalize_price_cache_key(get_row_value(row_values, product_code_index))
                weight_value = get_row_value(row_values, weight_index)
                cost_price_value = get_row_value(row_values, cost_price_index)
                cost_price = parse_number(cost_price_value)
                shipping_fee = calculate_shipping_fee(weight_value, shipping_fees)
                price = calculate_price(cost_price_value, shipping_fee, price_rules)

                if not product_code or price is None:
                    skipped_rows += 1
                elif product_code in price_cache:
                    price_cache[product_code] = {"cost_price": cost_price, "shipping_fee": shipping_fee, "price": price}
                    updated_rows += 1
                else:
                    price_cache[product_code] = {"cost_price": cost_price, "shipping_fee": shipping_fee, "price": price}
                    added_rows += 1

                if row_number % 500 == 0:
                    report_progress(
                        progress_callback,
                        f"正在导入{progress_label}：{source_sheet.title} 第 {row_number} 行",
                        processed_rows,
                        total_rows,
                    )

            summary.append((source_sheet.title, added_rows, updated_rows, skipped_rows))
            report_progress(
                progress_callback,
                f"已导入 {sheet_index}/{len(source.worksheets)} 个工作表",
                processed_rows,
                total_rows,
            )
    finally:
        source.close()

    return summary


def import_prices_to_cache_csv(
    input_path,
    progress_callback=None,
    shipping_fees=None,
    price_rules=None,
    price_cache=None,
    product_code_column=PRODUCT_CODE_COLUMN_NAME,
    weight_column=WEIGHT_COLUMN_NAME,
    cost_price_column=COST_PRICE_COLUMN_NAME,
    progress_label="价格库",
):
    price_cache = price_cache if price_cache is not None else {}

    with open(input_path, "r", newline="", encoding="utf-8-sig") as count_file:
        total_rows = max(sum(1 for _ in count_file) - 1, 0)

    with open(input_path, "r", newline="", encoding="utf-8-sig") as source_file:
        sample = source_file.read(4096)
        source_file.seek(0)
        dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
        reader = csv.reader(source_file, dialect)

        header = next(reader, None)
        if header is None:
            raise ValueError("CSV 文件为空。")

        product_code_index = find_header_index(header, product_code_column)
        weight_index = find_header_index(header, weight_column)
        cost_price_index = find_header_index(header, cost_price_column)

        if product_code_index == -1:
            raise ValueError(f"没有找到“{product_code_column}”列。")
        if weight_index == -1:
            raise ValueError(f"没有找到“{weight_column}”列。")
        if cost_price_index == -1:
            raise ValueError(f"没有找到“{cost_price_column}”列。")

        added_rows = 0
        updated_rows = 0
        skipped_rows = 0
        processed_rows = 0

        for row_number, row in enumerate(reader, start=2):
            processed_rows += 1
            product_code = normalize_price_cache_key(get_row_value(row, product_code_index))
            weight_value = get_row_value(row, weight_index)
            cost_price_value = get_row_value(row, cost_price_index)
            cost_price = parse_number(cost_price_value)
            shipping_fee = calculate_shipping_fee(weight_value, shipping_fees)
            price = calculate_price(cost_price_value, shipping_fee, price_rules)

            if not product_code or price is None:
                skipped_rows += 1
            elif product_code in price_cache:
                price_cache[product_code] = {"cost_price": cost_price, "shipping_fee": shipping_fee, "price": price}
                updated_rows += 1
            else:
                price_cache[product_code] = {"cost_price": cost_price, "shipping_fee": shipping_fee, "price": price}
                added_rows += 1

            if row_number % 500 == 0:
                report_progress(progress_callback, f"正在导入{progress_label} CSV 第 {row_number} 行", processed_rows, total_rows)

    report_progress(progress_callback, f"CSV {progress_label}导入完成。", total_rows, total_rows)

    return [("CSV", added_rows, updated_rows, skipped_rows)]


def import_sales_analysis_to_cache_xlsx(input_path, progress_callback=None, sales_cache=None, shipping_fees=None):
    source = load_workbook(input_path, read_only=True, data_only=True)
    sales_cache = sales_cache if sales_cache is not None else {}
    summary = []

    try:
        total_rows = estimate_xlsx_total_rows(source)
        processed_rows = 0

        for sheet_index, source_sheet in enumerate(source.worksheets, start=1):
            rows = source_sheet.iter_rows(values_only=True)
            header = next(rows, None)

            if header is None:
                summary.append((source_sheet.title, 0, 0, 0))
                continue

            headers = list(header)
            order_index = find_header_index(headers, INTERNAL_ORDER_NUMBER_COLUMN_NAME)
            weight_index = find_header_index(headers, ORDER_WEIGHT_COLUMN_NAME)
            amount_index = find_header_index(headers, SALES_AMOUNT_COLUMN_NAME)
            cost_index = find_header_index(headers, SALES_COST_COLUMN_NAME)

            if order_index == -1:
                raise ValueError(f"工作表“{source_sheet.title}”没有找到“{INTERNAL_ORDER_NUMBER_COLUMN_NAME}”列。")
            if weight_index == -1:
                raise ValueError(f"工作表“{source_sheet.title}”没有找到“{ORDER_WEIGHT_COLUMN_NAME}”列。")
            if amount_index == -1:
                raise ValueError(f"工作表“{source_sheet.title}”没有找到“{SALES_AMOUNT_COLUMN_NAME}”列。")
            if cost_index == -1:
                raise ValueError(f"工作表“{source_sheet.title}”没有找到“{SALES_COST_COLUMN_NAME}”列。")

            added_rows = 0
            updated_rows = 0
            skipped_rows = 0

            for row_number, row in enumerate(rows, start=2):
                row_values = list(row)
                processed_rows += 1
                order_number = str(get_row_value(row_values, order_index) or "").strip()

                if not order_number:
                    skipped_rows += 1
                else:
                    order_weight = parse_number(get_row_value(row_values, weight_index))
                    shipping_fee = calculate_shipping_fee(get_row_value(row_values, weight_index), shipping_fees)
                    sales_amount = parse_number(get_row_value(row_values, amount_index))
                    sales_cost = parse_number(get_row_value(row_values, cost_index))
                    record = {
                        "order_weight": order_weight,
                        "shipping_fee": shipping_fee,
                        "sales_amount": sales_amount,
                        "sales_cost": sales_cost,
                        "gross_profit": calculate_gross_profit(sales_amount, sales_cost, shipping_fee),
                    }
                    if order_number in sales_cache:
                        updated_rows += 1
                    else:
                        added_rows += 1
                    sales_cache[order_number] = record

                if row_number % 500 == 0:
                    report_progress(
                        progress_callback,
                        f"正在导入销售主题分析价格库：{source_sheet.title} 第 {row_number} 行",
                        processed_rows,
                        total_rows,
                    )

            summary.append((source_sheet.title, added_rows, updated_rows, skipped_rows))
            report_progress(
                progress_callback,
                f"已导入 {sheet_index}/{len(source.worksheets)} 个工作表",
                processed_rows,
                total_rows,
            )
    finally:
        source.close()

    return summary


def calculate_gross_profit(sales_amount, sales_cost, shipping_fee):
    shipping_fee_number = parse_number(shipping_fee)
    if sales_amount is None or sales_cost is None or shipping_fee_number is None:
        return None
    return round(sales_amount - sales_cost - shipping_fee_number, 2)


def import_sales_analysis_to_cache_csv(input_path, progress_callback=None, sales_cache=None, shipping_fees=None):
    sales_cache = sales_cache if sales_cache is not None else {}

    with open(input_path, "r", newline="", encoding="utf-8-sig") as count_file:
        total_rows = max(sum(1 for _ in count_file) - 1, 0)

    with open(input_path, "r", newline="", encoding="utf-8-sig") as source_file:
        sample = source_file.read(4096)
        source_file.seek(0)
        dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
        reader = csv.reader(source_file, dialect)

        header = next(reader, None)
        if header is None:
            raise ValueError("CSV 文件为空。")

        order_index = find_header_index(header, INTERNAL_ORDER_NUMBER_COLUMN_NAME)
        weight_index = find_header_index(header, ORDER_WEIGHT_COLUMN_NAME)
        amount_index = find_header_index(header, SALES_AMOUNT_COLUMN_NAME)
        cost_index = find_header_index(header, SALES_COST_COLUMN_NAME)

        if order_index == -1:
            raise ValueError(f"没有找到“{INTERNAL_ORDER_NUMBER_COLUMN_NAME}”列。")
        if weight_index == -1:
            raise ValueError(f"没有找到“{ORDER_WEIGHT_COLUMN_NAME}”列。")
        if amount_index == -1:
            raise ValueError(f"没有找到“{SALES_AMOUNT_COLUMN_NAME}”列。")
        if cost_index == -1:
            raise ValueError(f"没有找到“{SALES_COST_COLUMN_NAME}”列。")

        added_rows = 0
        updated_rows = 0
        skipped_rows = 0
        processed_rows = 0

        for row_number, row in enumerate(reader, start=2):
            processed_rows += 1
            order_number = str(get_row_value(row, order_index) or "").strip()

            if not order_number:
                skipped_rows += 1
            else:
                order_weight = parse_number(get_row_value(row, weight_index))
                shipping_fee = calculate_shipping_fee(get_row_value(row, weight_index), shipping_fees)
                sales_amount = parse_number(get_row_value(row, amount_index))
                sales_cost = parse_number(get_row_value(row, cost_index))
                record = {
                    "order_weight": order_weight,
                    "shipping_fee": shipping_fee,
                    "sales_amount": sales_amount,
                    "sales_cost": sales_cost,
                    "gross_profit": calculate_gross_profit(sales_amount, sales_cost, shipping_fee),
                }
                if order_number in sales_cache:
                    updated_rows += 1
                else:
                    added_rows += 1
                sales_cache[order_number] = record

            if row_number % 500 == 0:
                report_progress(progress_callback, f"正在导入销售主题分析价格库 CSV 第 {row_number} 行", processed_rows, total_rows)

    report_progress(progress_callback, "CSV 销售主题分析价格库导入完成。", total_rows, total_rows)

    return [("CSV", added_rows, updated_rows, skipped_rows)]


def default_output_path(input_path):
    folder = os.path.dirname(input_path)
    base, extension = os.path.splitext(os.path.basename(input_path))
    suffix = "_新增快递费列"

    if extension.lower() == ".csv":
        return os.path.join(folder, f"{base}{suffix}.csv")

    return os.path.join(folder, f"{base}{suffix}.xlsx")


def default_price_match_output_path(input_path):
    folder = os.path.dirname(input_path)
    base, extension = os.path.splitext(os.path.basename(input_path))
    suffix = "_匹配售价"

    if extension.lower() == ".csv":
        return os.path.join(folder, f"{base}{suffix}.csv")

    return os.path.join(folder, f"{base}{suffix}.xlsx")
