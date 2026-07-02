import csv
import json
import os
import re
import sys
import threading
import warnings
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook


warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    category=UserWarning,
    module="openpyxl.styles.stylesheet",
)


WEIGHT_COLUMN_NAME = "商品重量"
INSERTED_COLUMN_NAME = "快递费"
PRICE_COLUMN_NAME = "售价"
PRODUCT_COLUMN_NAME = "商品名"
PRODUCT_CODE_COLUMN_NAME = "商品编码"
COMBO_PRODUCT_CODE_COLUMN_NAME = "组合商品编码"
AVAILABLE_COLUMN_NAME = "可用数"
STOCK_COLUMN_NAME = "库存"
PRODUCT_TAG_COLUMN_NAME = "商品标签"
SUPPLIER_COLUMN_NAME = "供应商"
COST_PRICE_COLUMN_NAME = "成本价"
COMBO_WEIGHT_COLUMN_NAME = "组合重量"
COMBO_COST_PRICE_COLUMN_NAME = "组合成本价"
SALES_15_DAYS_COLUMN_NAME = "15天销量"
PRICE_CACHE_FILE_NAME = "price_cache.json"
DEFAULT_PRICE_MATCH_COLUMN_NAME = PRODUCT_CODE_COLUMN_NAME
DEFAULT_PRICE_THRESHOLD = 4.5
DEFAULT_LOW_PRICE_MARGIN = 0.13
DEFAULT_HIGH_PRICE_MARGIN = 0.13
SHIPPING_FEE_RULES = [
    ("0 < 商品重量 < 0.3", 1.3),
    ("0.3 <= 商品重量 < 0.5", 1.5),
    ("0.5 <= 商品重量 < 1", 1.9),
    ("1 <= 商品重量 < 1.5", 2.3),
    ("1.5 <= 商品重量 < 2", 2.8),
    ("2 <= 商品重量 < 2.5", 3.5),
    ("2.5 <= 商品重量 < 3", 3.9),
    ("3 <= 商品重量 < 4", 4.3),
    ("4 <= 商品重量 < 5", 5.6),
    ("商品重量 >= 5", 13.5),
]
DEFAULT_DELETE_RULES = [
    {
        "name": "商品名和成本价都空白",
        "conditions": [
            {"column": PRODUCT_COLUMN_NAME, "operator": "blank"},
            {"column": COST_PRICE_COLUMN_NAME, "operator": "blank"},
        ],
    },
    {
        "name": "可用数为0或空白且成本价空白",
        "conditions": [
            {"column": AVAILABLE_COLUMN_NAME, "operator": "zero_or_blank"},
            {"column": COST_PRICE_COLUMN_NAME, "operator": "blank"},
        ],
    },
    {
        "name": "商品名含下架且库存为0或空白",
        "conditions": [
            {"column": PRODUCT_COLUMN_NAME, "operator": "contains", "value": "下架"},
            {"column": STOCK_COLUMN_NAME, "operator": "zero_or_blank"},
        ],
    },
    {
        "name": "商品名含下架且可用数小于100",
        "conditions": [
            {"column": PRODUCT_COLUMN_NAME, "operator": "contains", "value": "下架"},
            {"column": AVAILABLE_COLUMN_NAME, "operator": "lt", "value": 100},
        ],
    },
    {
        "name": "商品名含zxp且可用数小于100",
        "conditions": [
            {"column": PRODUCT_COLUMN_NAME, "operator": "contains_ci", "value": "zxp"},
            {"column": AVAILABLE_COLUMN_NAME, "operator": "lt", "value": 100},
        ],
    },
    {
        "name": "商品名含微瑕品",
        "conditions": [
            {"column": PRODUCT_COLUMN_NAME, "operator": "contains", "value": "微瑕品"},
        ],
    },
    {
        "name": "可用数小于10且商品重量空白",
        "conditions": [
            {"column": AVAILABLE_COLUMN_NAME, "operator": "lt", "value": 10},
            {"column": WEIGHT_COLUMN_NAME, "operator": "blank"},
        ],
    },
    {
        "name": "可用数为0且供应商没有英文字母",
        "conditions": [
            {"column": AVAILABLE_COLUMN_NAME, "operator": "eq", "value": 0},
            {"column": SUPPLIER_COLUMN_NAME, "operator": "no_english_letter"},
        ],
    },
    {
        "name": "商品标签空白且可用数为0",
        "conditions": [
            {"column": PRODUCT_TAG_COLUMN_NAME, "operator": "blank"},
            {"column": AVAILABLE_COLUMN_NAME, "operator": "eq", "value": 0},
        ],
    },
    {
        "name": "商品标签含家具或赠品",
        "any": [
            {"column": PRODUCT_TAG_COLUMN_NAME, "operator": "contains", "value": "家具"},
            {"column": PRODUCT_TAG_COLUMN_NAME, "operator": "contains", "value": "赠品"},
        ],
    },
    {
        "name": "商品标签含待清理且可用数小于30",
        "conditions": [
            {"column": PRODUCT_TAG_COLUMN_NAME, "operator": "contains", "value": "待清理"},
            {"column": AVAILABLE_COLUMN_NAME, "operator": "lt", "value": 30},
        ],
    },
    {
        "name": "商品标签含分销商且可用数小于100",
        "conditions": [
            {"column": PRODUCT_TAG_COLUMN_NAME, "operator": "contains", "value": "分销商"},
            {"column": AVAILABLE_COLUMN_NAME, "operator": "lt", "value": 100},
        ],
    },
    {
        "name": "商品标签4位数字小于2601且可用数小于20且15天销量小于15",
        "conditions": [
            {"column": PRODUCT_TAG_COLUMN_NAME, "operator": "four_digit_lt", "value": 2601},
            {"column": AVAILABLE_COLUMN_NAME, "operator": "lt", "value": 20},
            {"column": SALES_15_DAYS_COLUMN_NAME, "operator": "lt", "value": 15},
        ],
    },
    {
        "name": "供应商和成本价都空白",
        "conditions": [
            {"column": SUPPLIER_COLUMN_NAME, "operator": "blank"},
            {"column": COST_PRICE_COLUMN_NAME, "operator": "blank"},
        ],
    },
]


def is_blank(value):
    return str(value if value is not None else "").strip() == ""


def is_zero_or_blank(value):
    if is_blank(value):
        return True

    try:
        return float(str(value).strip().replace(",", "")) == 0
    except ValueError:
        return False


def parse_number(value):
    if is_blank(value):
        return None

    try:
        return float(str(value).strip().replace(",", ""))
    except ValueError:
        return None


def get_row_value(row, index):
    return row[index] if index != -1 and index < len(row) else None


def text_contains(value, text):
    return text in str(value if value is not None else "")


def text_contains_case_insensitive(value, text):
    return text.lower() in str(value if value is not None else "").lower()


def contains_english_letter(value):
    return any("A" <= char <= "Z" or "a" <= char <= "z" for char in str(value if value is not None else ""))


def has_four_digit_number_less_than(value, threshold):
    text = str(value if value is not None else "")
    return any(int(match) < threshold for match in re.findall(r"\d{4}", text))


def find_header_index(headers, name):
    normalized = [str(value if value is not None else "").strip() for value in headers]

    for index, header in enumerate(normalized):
        if header == name:
            return index

    for index, header in enumerate(normalized):
        if name in header:
            return index

    return -1


def app_base_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def resource_path(filename):
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, filename)
    return os.path.join(app_base_dir(), filename)


def default_rules_path():
    return os.path.join(app_base_dir(), "rules.json")


def default_price_cache_path():
    data_root = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
    data_dir = os.path.join(data_root, "图书电商线上活动价格自动生成器")
    os.makedirs(data_dir, exist_ok=True)
    return os.path.join(data_dir, PRICE_CACHE_FILE_NAME)


def normalize_match_key(value):
    return str(value if value is not None else "").strip()


def load_price_cache(path=None):
    cache_path = path or default_price_cache_path()

    if not cache_path or not os.path.exists(cache_path):
        return {}

    with open(cache_path, "r", encoding="utf-8") as cache_file:
        data = json.load(cache_file)

    prices = data.get("prices", data if isinstance(data, dict) else None)
    if not isinstance(prices, dict):
        raise ValueError("售价保存文件格式错误：需要 prices 对象。")

    return {normalize_match_key(key): value for key, value in prices.items() if normalize_match_key(key)}


def save_price_cache(price_cache, path=None):
    cache_path = path or default_price_cache_path()

    with open(cache_path, "w", encoding="utf-8") as cache_file:
        json.dump({"prices": price_cache}, cache_file, ensure_ascii=False, indent=2)


def load_delete_rules(path=None):
    rules_path = path or default_rules_path()

    if not rules_path or not os.path.exists(rules_path):
        return DEFAULT_DELETE_RULES

    with open(rules_path, "r", encoding="utf-8") as rules_file:
        data = json.load(rules_file)

    rules = data.get("delete_rules", data if isinstance(data, list) else None)
    if not isinstance(rules, list):
        raise ValueError("规则文件格式错误：需要 delete_rules 列表。")

    return rules


def save_default_rules(path):
    with open(path, "w", encoding="utf-8") as rules_file:
        json.dump({"delete_rules": DEFAULT_DELETE_RULES}, rules_file, ensure_ascii=False, indent=2)


def build_header_map(headers):
    return {str(value if value is not None else "").strip(): index for index, value in enumerate(headers)}


def get_row_column_value(row, header_map, column_name):
    index = header_map.get(column_name)
    if index is None:
        for header, header_index in header_map.items():
            if column_name in header:
                index = header_index
                break

    return get_row_value(row, index) if index is not None else None


def compare_number(value, operator, expected):
    number = parse_number(value)
    if number is None:
        return False

    if operator == "lt":
        return number < expected
    if operator == "lte":
        return number <= expected
    if operator == "gt":
        return number > expected
    if operator == "gte":
        return number >= expected
    if operator == "eq":
        return number == expected

    raise ValueError(f"不支持的数字比较：{operator}")


def evaluate_condition(row, header_map, condition):
    column_name = condition.get("column")
    operator = condition.get("operator")
    expected = condition.get("value")
    value = get_row_column_value(row, header_map, column_name)

    if operator == "blank":
        return is_blank(value)
    if operator == "not_blank":
        return not is_blank(value)
    if operator == "zero_or_blank":
        return is_zero_or_blank(value)
    if operator == "contains":
        return text_contains(value, str(expected))
    if operator == "contains_ci":
        return text_contains_case_insensitive(value, str(expected))
    if operator == "no_english_letter":
        return not contains_english_letter(value)
    if operator == "four_digit_lt":
        return has_four_digit_number_less_than(value, float(expected))
    if operator in {"lt", "lte", "gt", "gte", "eq"}:
        return compare_number(value, operator, float(expected))

    raise ValueError(f"不支持的规则操作：{operator}")


def should_delete_row(row, header_map, delete_rules):
    for rule in delete_rules:
        conditions = rule.get("conditions", [])
        any_conditions = rule.get("any", [])

        if conditions and all(evaluate_condition(row, header_map, condition) for condition in conditions):
            return True
        if any_conditions and any(evaluate_condition(row, header_map, condition) for condition in any_conditions):
            return True

    return False


def parse_weight(value):
    if value is None:
        return None

    text = str(value).strip().replace(",", "")
    number = ""
    dot_seen = False

    for char in text:
        if char.isdigit():
            number += char
        elif char == "." and not dot_seen:
            number += char
            dot_seen = True
        elif number:
            break

    if not number or number == ".":
        return None

    try:
        return float(number)
    except ValueError:
        return None


def calculate_shipping_fee(value, shipping_fees=None):
    weight = parse_weight(value)
    fees = shipping_fees or [rule[1] for rule in SHIPPING_FEE_RULES]

    if weight is None or weight == 0:
        return "请检查"
    if 0 < weight < 0.3:
        return fees[0]
    if 0.3 <= weight < 0.5:
        return fees[1]
    if 0.5 <= weight < 1:
        return fees[2]
    if 1 <= weight < 1.5:
        return fees[3]
    if 1.5 <= weight < 2:
        return fees[4]
    if 2 <= weight < 2.5:
        return fees[5]
    if 2.5 <= weight < 3:
        return fees[6]
    if 3 <= weight < 4:
        return fees[7]
    if 4 <= weight < 5:
        return fees[8]
    if weight >= 5:
        return fees[9]

    return "请检查"


def insert_value(row, index, value):
    next_row = list(row)
    while len(next_row) < index:
        next_row.append(None)
    next_row.insert(index, value)
    return next_row


def calculate_price(cost_price_value, shipping_fee_value, price_rules=None):
    cost_price = parse_number(cost_price_value)
    shipping_fee = parse_number(shipping_fee_value)

    if cost_price is None or shipping_fee is None:
        return None

    threshold, low_margin, high_margin = price_rules or (
        DEFAULT_PRICE_THRESHOLD,
        DEFAULT_LOW_PRICE_MARGIN,
        DEFAULT_HIGH_PRICE_MARGIN,
    )
    margin = high_margin if cost_price > threshold else low_margin
    divisor = 1 - margin
    return round((cost_price + shipping_fee) / divisor, 2)


def resolve_price(row, header_map, cost_price_index, shipping_fee, price_rules, price_match_column, price_cache):
    match_key = normalize_match_key(get_row_column_value(row, header_map, price_match_column))

    if match_key and match_key in price_cache:
        return price_cache[match_key], True, False

    cost_price_value = get_row_value(row, cost_price_index)
    price = calculate_price(cost_price_value, shipping_fee, price_rules)

    if match_key and price is not None:
        price_cache[match_key] = price
        return price, False, True

    return price, False, False


def insert_values(row, index, values):
    next_row = list(row)
    while len(next_row) < index:
        next_row.append(None)

    for offset, value in enumerate(values):
        next_row.insert(index + offset, value)

    return next_row


def report_progress(progress_callback, text, current=None, total=None):
    if not progress_callback:
        return

    percent = None
    if current is not None and total:
        percent = max(0, min(100, int(current * 100 / total)))

    progress_callback(text, percent)


def estimate_xlsx_total_rows(workbook):
    total = 0

    for sheet in workbook.worksheets:
        max_row = sheet.max_row or 0
        if max_row > 1:
            total += max_row - 1

    return total or None


def process_xlsx(
    input_path,
    output_path,
    insert_side="after",
    progress_callback=None,
    shipping_fees=None,
    price_rules=None,
    delete_rules=None,
    price_match_column=DEFAULT_PRICE_MATCH_COLUMN_NAME,
    price_cache=None,
):
    source = load_workbook(input_path, read_only=True, data_only=True)
    target = Workbook(write_only=True)
    price_cache = price_cache if price_cache is not None else {}

    if target.worksheets:
        target.remove(target.worksheets[0])

    summary = []

    try:
        total_rows = estimate_xlsx_total_rows(source)
        processed_rows = 0

        for sheet_index, source_sheet in enumerate(source.worksheets, start=1):
            target_sheet = target.create_sheet(title=source_sheet.title)
            rows = source_sheet.iter_rows(values_only=True)
            header = next(rows, None)

            if header is None:
                target_sheet.append([])
                summary.append((source_sheet.title, 0, 0, 0, 0))
                continue

            headers = list(header)
            header_map = build_header_map(headers)
            weight_index = find_header_index(headers, WEIGHT_COLUMN_NAME)
            product_index = find_header_index(headers, PRODUCT_COLUMN_NAME)
            cost_price_index = find_header_index(headers, COST_PRICE_COLUMN_NAME)
            available_index = find_header_index(headers, AVAILABLE_COLUMN_NAME)
            stock_index = find_header_index(headers, STOCK_COLUMN_NAME)
            product_tag_index = find_header_index(headers, PRODUCT_TAG_COLUMN_NAME)
            supplier_index = find_header_index(headers, SUPPLIER_COLUMN_NAME)
            sales_15_days_index = find_header_index(headers, SALES_15_DAYS_COLUMN_NAME)

            if weight_index == -1:
                raise ValueError(f"工作表“{source_sheet.title}”没有找到“{WEIGHT_COLUMN_NAME}”列。")
            if find_header_index(headers, price_match_column) == -1:
                raise ValueError(f"工作表“{source_sheet.title}”没有找到售价匹配列“{price_match_column}”。")

            insert_index = weight_index if insert_side == "before" else weight_index + 1
            target_sheet.append(insert_values(headers, insert_index, [INSERTED_COLUMN_NAME, PRICE_COLUMN_NAME]))

            kept_rows = 0
            deleted_rows = 0
            matched_prices = 0
            saved_prices = 0

            for row_number, row in enumerate(rows, start=2):
                row_values = list(row)
                processed_rows += 1

                if should_delete_row(row_values, header_map, delete_rules or DEFAULT_DELETE_RULES):
                    deleted_rows += 1
                    if row_number % 500 == 0:
                        report_progress(
                            progress_callback,
                            f"正在处理：{source_sheet.title} 第 {row_number} 行",
                            processed_rows,
                            total_rows,
                        )
                    continue

                weight_value = get_row_value(row_values, weight_index)
                shipping_fee = calculate_shipping_fee(weight_value, shipping_fees)
                price, matched, saved = resolve_price(
                    row_values,
                    header_map,
                    cost_price_index,
                    shipping_fee,
                    price_rules,
                    price_match_column,
                    price_cache,
                )
                if matched:
                    matched_prices += 1
                if saved:
                    saved_prices += 1
                target_sheet.append(insert_values(row_values, insert_index, [shipping_fee, price]))
                kept_rows += 1

                if row_number % 500 == 0:
                    report_progress(
                        progress_callback,
                        f"正在处理：{source_sheet.title} 第 {row_number} 行",
                        processed_rows,
                        total_rows,
                    )

            summary.append((source_sheet.title, kept_rows, deleted_rows, matched_prices, saved_prices))

            report_progress(
                progress_callback,
                f"已处理 {sheet_index}/{len(source.worksheets)} 个工作表",
                processed_rows,
                total_rows,
            )

        report_progress(progress_callback, "正在保存导出文件...", total_rows, total_rows)
        target.save(output_path)
    finally:
        source.close()
        target.close()

    return summary


def process_csv(
    input_path,
    output_path,
    insert_side="after",
    progress_callback=None,
    shipping_fees=None,
    price_rules=None,
    delete_rules=None,
    price_match_column=DEFAULT_PRICE_MATCH_COLUMN_NAME,
    price_cache=None,
):
    price_cache = price_cache if price_cache is not None else {}

    with open(input_path, "r", newline="", encoding="utf-8-sig") as count_file:
        total_rows = max(sum(1 for _ in count_file) - 1, 0)

    with open(input_path, "r", newline="", encoding="utf-8-sig") as source_file:
        sample = source_file.read(4096)
        source_file.seek(0)
        dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
        reader = csv.reader(source_file, dialect)

        header = next(reader, None)
        if header is None:
            raise ValueError("CSV 文件为空。")

        header_map = build_header_map(header)
        weight_index = find_header_index(header, WEIGHT_COLUMN_NAME)
        product_index = find_header_index(header, PRODUCT_COLUMN_NAME)
        cost_price_index = find_header_index(header, COST_PRICE_COLUMN_NAME)
        available_index = find_header_index(header, AVAILABLE_COLUMN_NAME)
        stock_index = find_header_index(header, STOCK_COLUMN_NAME)
        product_tag_index = find_header_index(header, PRODUCT_TAG_COLUMN_NAME)
        supplier_index = find_header_index(header, SUPPLIER_COLUMN_NAME)
        sales_15_days_index = find_header_index(header, SALES_15_DAYS_COLUMN_NAME)

        if weight_index == -1:
            raise ValueError(f"没有找到“{WEIGHT_COLUMN_NAME}”列。")
        if find_header_index(header, price_match_column) == -1:
            raise ValueError(f"没有找到售价匹配列“{price_match_column}”。")

        insert_index = weight_index if insert_side == "before" else weight_index + 1

        with open(output_path, "w", newline="", encoding="utf-8-sig") as target_file:
            writer = csv.writer(target_file)
            writer.writerow(insert_values(header, insert_index, [INSERTED_COLUMN_NAME, PRICE_COLUMN_NAME]))

            kept_rows = 0
            deleted_rows = 0
            matched_prices = 0
            saved_prices = 0
            processed_rows = 0

            for row_number, row in enumerate(reader, start=2):
                processed_rows += 1

                if should_delete_row(row, header_map, delete_rules or DEFAULT_DELETE_RULES):
                    deleted_rows += 1
                    if row_number % 500 == 0:
                        report_progress(progress_callback, f"正在处理 CSV 第 {row_number} 行", processed_rows, total_rows)
                    continue

                weight_value = get_row_value(row, weight_index)
                shipping_fee = calculate_shipping_fee(weight_value, shipping_fees)
                price, matched, saved = resolve_price(
                    row,
                    header_map,
                    cost_price_index,
                    shipping_fee,
                    price_rules,
                    price_match_column,
                    price_cache,
                )
                if matched:
                    matched_prices += 1
                if saved:
                    saved_prices += 1
                writer.writerow(insert_values(row, insert_index, [shipping_fee, price]))
                kept_rows += 1

                if row_number % 500 == 0:
                    report_progress(progress_callback, f"正在处理 CSV 第 {row_number} 行", processed_rows, total_rows)

    report_progress(progress_callback, "CSV 处理完成。", total_rows, total_rows)

    return [("CSV", kept_rows, deleted_rows, matched_prices, saved_prices)]


def append_cached_prices_xlsx(input_path, output_path, progress_callback=None, price_cache=None):
    source = load_workbook(input_path, read_only=True, data_only=True)
    target = Workbook(write_only=True)
    price_cache = price_cache if price_cache is not None else {}

    if target.worksheets:
        target.remove(target.worksheets[0])

    summary = []

    try:
        total_rows = estimate_xlsx_total_rows(source)
        processed_rows = 0

        for sheet_index, source_sheet in enumerate(source.worksheets, start=1):
            target_sheet = target.create_sheet(title=source_sheet.title)
            rows = source_sheet.iter_rows(values_only=True)
            header = next(rows, None)

            if header is None:
                target_sheet.append([PRICE_COLUMN_NAME])
                summary.append((source_sheet.title, 0, 0))
                continue

            headers = list(header)
            header_map = build_header_map(headers)

            if find_header_index(headers, PRODUCT_CODE_COLUMN_NAME) == -1:
                raise ValueError(f"工作表“{source_sheet.title}”没有找到“{PRODUCT_CODE_COLUMN_NAME}”列。")

            target_sheet.append(headers + [PRICE_COLUMN_NAME])

            matched_rows = 0
            unmatched_rows = 0

            for row_number, row in enumerate(rows, start=2):
                row_values = list(row)
                processed_rows += 1
                product_code = normalize_match_key(get_row_column_value(row_values, header_map, PRODUCT_CODE_COLUMN_NAME))
                price = price_cache.get(product_code) if product_code else None

                if price is None:
                    unmatched_rows += 1
                else:
                    matched_rows += 1

                target_sheet.append(row_values + [price])

                if row_number % 500 == 0:
                    report_progress(
                        progress_callback,
                        f"正在匹配售价：{source_sheet.title} 第 {row_number} 行",
                        processed_rows,
                        total_rows,
                    )

            summary.append((source_sheet.title, matched_rows, unmatched_rows))
            report_progress(
                progress_callback,
                f"已匹配 {sheet_index}/{len(source.worksheets)} 个工作表",
                processed_rows,
                total_rows,
            )

        report_progress(progress_callback, "正在保存匹配结果...", total_rows, total_rows)
        target.save(output_path)
    finally:
        source.close()
        target.close()

    return summary


def append_cached_prices_csv(input_path, output_path, progress_callback=None, price_cache=None):
    price_cache = price_cache if price_cache is not None else {}

    with open(input_path, "r", newline="", encoding="utf-8-sig") as count_file:
        total_rows = max(sum(1 for _ in count_file) - 1, 0)

    with open(input_path, "r", newline="", encoding="utf-8-sig") as source_file:
        sample = source_file.read(4096)
        source_file.seek(0)
        dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
        reader = csv.reader(source_file, dialect)

        header = next(reader, None)
        if header is None:
            raise ValueError("CSV 文件为空。")

        header_map = build_header_map(header)
        if find_header_index(header, PRODUCT_CODE_COLUMN_NAME) == -1:
            raise ValueError(f"没有找到“{PRODUCT_CODE_COLUMN_NAME}”列。")

        with open(output_path, "w", newline="", encoding="utf-8-sig") as target_file:
            writer = csv.writer(target_file)
            writer.writerow(list(header) + [PRICE_COLUMN_NAME])

            matched_rows = 0
            unmatched_rows = 0
            processed_rows = 0

            for row_number, row in enumerate(reader, start=2):
                processed_rows += 1
                product_code = normalize_match_key(get_row_column_value(row, header_map, PRODUCT_CODE_COLUMN_NAME))
                price = price_cache.get(product_code) if product_code else None

                if price is None:
                    unmatched_rows += 1
                else:
                    matched_rows += 1

                writer.writerow(list(row) + [price])

                if row_number % 500 == 0:
                    report_progress(progress_callback, f"正在匹配 CSV 第 {row_number} 行", processed_rows, total_rows)

    report_progress(progress_callback, "CSV 售价匹配完成。", total_rows, total_rows)

    return [("CSV", matched_rows, unmatched_rows)]


def import_prices_to_cache_xlsx(
    input_path,
    progress_callback=None,
    shipping_fees=None,
    price_rules=None,
    price_cache=None,
    product_code_column=PRODUCT_CODE_COLUMN_NAME,
    weight_column=WEIGHT_COLUMN_NAME,
    cost_price_column=COST_PRICE_COLUMN_NAME,
    progress_label="价格库",
):
    source = load_workbook(input_path, read_only=True, data_only=True)
    price_cache = price_cache if price_cache is not None else {}
    summary = []

    try:
        total_rows = estimate_xlsx_total_rows(source)
        processed_rows = 0

        for sheet_index, source_sheet in enumerate(source.worksheets, start=1):
            rows = source_sheet.iter_rows(values_only=True)
            header = next(rows, None)

            if header is None:
                summary.append((source_sheet.title, 0, 0, 0))
                continue

            headers = list(header)
            header_map = build_header_map(headers)
            product_code_index = find_header_index(headers, product_code_column)
            weight_index = find_header_index(headers, weight_column)
            cost_price_index = find_header_index(headers, cost_price_column)

            if product_code_index == -1:
                raise ValueError(f"工作表“{source_sheet.title}”没有找到“{product_code_column}”列。")
            if weight_index == -1:
                raise ValueError(f"工作表“{source_sheet.title}”没有找到“{weight_column}”列。")
            if cost_price_index == -1:
                raise ValueError(f"工作表“{source_sheet.title}”没有找到“{cost_price_column}”列。")

            added_rows = 0
            updated_rows = 0
            skipped_rows = 0

            for row_number, row in enumerate(rows, start=2):
                row_values = list(row)
                processed_rows += 1
                product_code = normalize_match_key(get_row_value(row_values, product_code_index))
                weight_value = get_row_value(row_values, weight_index)
                cost_price_value = get_row_value(row_values, cost_price_index)
                shipping_fee = calculate_shipping_fee(weight_value, shipping_fees)
                price = calculate_price(cost_price_value, shipping_fee, price_rules)

                if not product_code or price is None:
                    skipped_rows += 1
                elif product_code in price_cache:
                    price_cache[product_code] = price
                    updated_rows += 1
                else:
                    price_cache[product_code] = price
                    added_rows += 1

                if row_number % 500 == 0:
                    report_progress(
                        progress_callback,
                        f"正在导入{progress_label}：{source_sheet.title} 第 {row_number} 行",
                        processed_rows,
                        total_rows,
                    )

            summary.append((source_sheet.title, added_rows, updated_rows, skipped_rows))
            report_progress(
                progress_callback,
                f"已导入 {sheet_index}/{len(source.worksheets)} 个工作表",
                processed_rows,
                total_rows,
            )
    finally:
        source.close()

    return summary


def import_prices_to_cache_csv(
    input_path,
    progress_callback=None,
    shipping_fees=None,
    price_rules=None,
    price_cache=None,
    product_code_column=PRODUCT_CODE_COLUMN_NAME,
    weight_column=WEIGHT_COLUMN_NAME,
    cost_price_column=COST_PRICE_COLUMN_NAME,
    progress_label="价格库",
):
    price_cache = price_cache if price_cache is not None else {}

    with open(input_path, "r", newline="", encoding="utf-8-sig") as count_file:
        total_rows = max(sum(1 for _ in count_file) - 1, 0)

    with open(input_path, "r", newline="", encoding="utf-8-sig") as source_file:
        sample = source_file.read(4096)
        source_file.seek(0)
        dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
        reader = csv.reader(source_file, dialect)

        header = next(reader, None)
        if header is None:
            raise ValueError("CSV 文件为空。")

        product_code_index = find_header_index(header, product_code_column)
        weight_index = find_header_index(header, weight_column)
        cost_price_index = find_header_index(header, cost_price_column)

        if product_code_index == -1:
            raise ValueError(f"没有找到“{product_code_column}”列。")
        if weight_index == -1:
            raise ValueError(f"没有找到“{weight_column}”列。")
        if cost_price_index == -1:
            raise ValueError(f"没有找到“{cost_price_column}”列。")

        added_rows = 0
        updated_rows = 0
        skipped_rows = 0
        processed_rows = 0

        for row_number, row in enumerate(reader, start=2):
            processed_rows += 1
            product_code = normalize_match_key(get_row_value(row, product_code_index))
            weight_value = get_row_value(row, weight_index)
            cost_price_value = get_row_value(row, cost_price_index)
            shipping_fee = calculate_shipping_fee(weight_value, shipping_fees)
            price = calculate_price(cost_price_value, shipping_fee, price_rules)

            if not product_code or price is None:
                skipped_rows += 1
            elif product_code in price_cache:
                price_cache[product_code] = price
                updated_rows += 1
            else:
                price_cache[product_code] = price
                added_rows += 1

            if row_number % 500 == 0:
                report_progress(progress_callback, f"正在导入{progress_label} CSV 第 {row_number} 行", processed_rows, total_rows)

    report_progress(progress_callback, f"CSV {progress_label}导入完成。", total_rows, total_rows)

    return [("CSV", added_rows, updated_rows, skipped_rows)]


def default_output_path(input_path):
    folder = os.path.dirname(input_path)
    base, extension = os.path.splitext(os.path.basename(input_path))
    suffix = "_新增快递费列"

    if extension.lower() == ".csv":
        return os.path.join(folder, f"{base}{suffix}.csv")

    return os.path.join(folder, f"{base}{suffix}.xlsx")


def default_price_match_output_path(input_path):
    folder = os.path.dirname(input_path)
    base, extension = os.path.splitext(os.path.basename(input_path))
    suffix = "_匹配售价"

    if extension.lower() == ".csv":
        return os.path.join(folder, f"{base}{suffix}.csv")

    return os.path.join(folder, f"{base}{suffix}.xlsx")


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("图书电商线上活动价格自动生成器")
        icon_path = resource_path("app.ico")
        if os.path.exists(icon_path):
            self.iconbitmap(icon_path)
        self.geometry("780x690")
        self.minsize(720, 640)
        self.input_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.cache_import_path = tk.StringVar()
        self.combo_cache_import_path = tk.StringVar()
        self.match_input_path = tk.StringVar()
        self.match_output_path = tk.StringVar()
        self.rules_path = tk.StringVar(value=default_rules_path())
        self.status = tk.StringVar(value="请选择表格导入价格库，或从价格库匹配售价。")
        self.progress = tk.IntVar(value=0)
        self.shipping_fee_vars = [tk.StringVar(value=str(default)) for _, default in SHIPPING_FEE_RULES]
        self.price_threshold = tk.StringVar(value=str(DEFAULT_PRICE_THRESHOLD))
        self.low_price_margin = tk.StringVar(value=str(DEFAULT_LOW_PRICE_MARGIN))
        self.high_price_margin = tk.StringVar(value=str(DEFAULT_HIGH_PRICE_MARGIN))
        self.combo_price_threshold = tk.StringVar(value=str(DEFAULT_PRICE_THRESHOLD))
        self.combo_low_price_margin = tk.StringVar(value=str(DEFAULT_LOW_PRICE_MARGIN))
        self.combo_high_price_margin = tk.StringVar(value=str(DEFAULT_HIGH_PRICE_MARGIN))
        self.price_match_column = tk.StringVar(value=DEFAULT_PRICE_MATCH_COLUMN_NAME)
        self._build_ui()

    def _build_ui(self):
        frame = ttk.Frame(self, padding=18)
        frame.pack(fill="both", expand=True)
        frame.columnconfigure(1, weight=1)

        price_frames = ttk.Frame(frame)
        price_frames.grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 12))
        price_frames.columnconfigure(0, weight=1)
        price_frames.columnconfigure(1, weight=1)

        price_frame = ttk.LabelFrame(price_frames, text="普通商品售价规则", padding=10)
        price_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        ttk.Label(price_frame, text="分界成本价").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(price_frame, textvariable=self.price_threshold, width=8).grid(row=0, column=1, sticky="w", pady=3)
        ttk.Label(price_frame, text="成本<=分界 毛利率").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(price_frame, textvariable=self.low_price_margin, width=8).grid(row=1, column=1, sticky="w", pady=3)
        ttk.Label(price_frame, text="成本>分界 毛利率").grid(row=1, column=2, sticky="w", padx=(16, 8), pady=3)
        ttk.Entry(price_frame, textvariable=self.high_price_margin, width=8).grid(row=1, column=3, sticky="w", pady=3)

        combo_price_frame = ttk.LabelFrame(price_frames, text="严选组合售价规则", padding=10)
        combo_price_frame.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
        ttk.Label(combo_price_frame, text="分界成本价").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(combo_price_frame, textvariable=self.combo_price_threshold, width=8).grid(row=0, column=1, sticky="w", pady=3)
        ttk.Label(combo_price_frame, text="成本<=分界 毛利率").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(combo_price_frame, textvariable=self.combo_low_price_margin, width=8).grid(row=1, column=1, sticky="w", pady=3)
        ttk.Label(combo_price_frame, text="成本>分界 毛利率").grid(row=1, column=2, sticky="w", padx=(16, 8), pady=3)
        ttk.Entry(combo_price_frame, textvariable=self.combo_high_price_margin, width=8).grid(row=1, column=3, sticky="w", pady=3)

        fee_frame = ttk.LabelFrame(frame, text="快递费规则（填写返回数字）", padding=10)
        fee_frame.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(0, 12))
        fee_frame.columnconfigure(1, weight=1)
        fee_frame.columnconfigure(3, weight=1)

        for index, (label, _) in enumerate(SHIPPING_FEE_RULES):
            row = index // 2
            column = (index % 2) * 2
            ttk.Label(fee_frame, text=label).grid(row=row, column=column, sticky="w", padx=(0, 8), pady=3)
            ttk.Entry(fee_frame, textvariable=self.shipping_fee_vars[index], width=8).grid(
                row=row, column=column + 1, sticky="w", padx=(0, 18), pady=3
            )

        import_frames = ttk.Frame(frame)
        import_frames.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(0, 12))
        import_frames.columnconfigure(0, weight=1)
        import_frames.columnconfigure(1, weight=1)

        import_cache_frame = ttk.LabelFrame(import_frames, text="导入表格到价格库", padding=10)
        import_cache_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        ttk.Label(import_cache_frame, text="导入表格").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(import_cache_frame, textvariable=self.cache_import_path, width=28).grid(row=0, column=1, sticky="w", padx=(0, 8), pady=3)
        ttk.Button(import_cache_frame, text="选择", command=self.choose_cache_import).grid(row=0, column=2, sticky="ew", pady=3)
        ttk.Label(
            import_cache_frame,
            text="读取“商品编码”“商品重量”“成本价”，按当前公式计算售价并保存到价格库。",
            foreground="#435064",
            wraplength=320,
        ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(6, 8))
        self.cache_import_button = ttk.Button(import_cache_frame, text="计算并导入价格库", command=self.start_price_cache_import)
        self.cache_import_button.grid(row=2, column=0, columnspan=3, sticky="ew", ipady=6)

        combo_cache_frame = ttk.LabelFrame(import_frames, text="导入严选组合商品编码到价格库", padding=10)
        combo_cache_frame.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
        ttk.Label(combo_cache_frame, text="导入表格").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(combo_cache_frame, textvariable=self.combo_cache_import_path, width=28).grid(row=0, column=1, sticky="w", padx=(0, 8), pady=3)
        ttk.Button(combo_cache_frame, text="选择", command=self.choose_combo_cache_import).grid(row=0, column=2, sticky="ew", pady=3)
        ttk.Label(
            combo_cache_frame,
            text="读取“组合商品编码”“组合重量”“组合成本价”，按当前公式计算售价并保存到价格库。",
            foreground="#435064",
            wraplength=320,
        ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(6, 8))
        self.combo_cache_import_button = ttk.Button(
            combo_cache_frame,
            text="计算并导入价格库",
            command=self.start_combo_price_cache_import,
        )
        self.combo_cache_import_button.grid(row=2, column=0, columnspan=3, sticky="ew", ipady=6)

        cache_frame = ttk.LabelFrame(frame, text="从价格库匹配售价", padding=10)
        cache_frame.grid(row=3, column=0, columnspan=3, sticky="ew", pady=(0, 12))
        cache_frame.columnconfigure(1, weight=1)
        ttk.Label(cache_frame, text="导入新表格").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(cache_frame, textvariable=self.match_input_path).grid(row=0, column=1, sticky="ew", padx=(0, 8), pady=3)
        ttk.Button(cache_frame, text="选择", command=self.choose_match_input).grid(row=0, column=2, sticky="ew", pady=3)
        ttk.Label(cache_frame, text="导出新表格").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(cache_frame, textvariable=self.match_output_path).grid(row=1, column=1, sticky="ew", padx=(0, 8), pady=3)
        ttk.Button(cache_frame, text="另存为", command=self.choose_match_output).grid(row=1, column=2, sticky="ew", pady=3)
        ttk.Label(cache_frame, text="固定查找“商品编码”，在新表格末尾新增“售价”列。", foreground="#435064").grid(
            row=2, column=0, columnspan=3, sticky="w", pady=(6, 8)
        )
        self.match_button = ttk.Button(cache_frame, text="匹配售价并导出", command=self.start_price_match)
        self.match_button.grid(row=3, column=0, columnspan=3, sticky="ew", ipady=6)

        self.progress_bar = ttk.Progressbar(frame, variable=self.progress, maximum=100, mode="determinate")
        self.progress_bar.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(14, 0))

        ttk.Label(frame, textvariable=self.status, foreground="#1f7a8c", wraplength=590).grid(
            row=5, column=0, columnspan=3, sticky="w", pady=(10, 0)
        )

    def choose_input(self):
        path = filedialog.askopenfilename(
            title="选择表格文件",
            filetypes=[
                ("表格文件", "*.xlsx *.xlsm *.csv"),
                ("Excel 文件", "*.xlsx *.xlsm"),
                ("CSV 文件", "*.csv"),
                ("所有文件", "*.*"),
            ],
        )

        if not path:
            return

        self.input_path.set(path)
        self.output_path.set(default_output_path(path))

    def choose_output(self):
        initial = self.output_path.get() or default_output_path(self.input_path.get() or "导出.xlsx")
        extension = os.path.splitext(initial)[1].lower()
        filetypes = [("Excel 文件", "*.xlsx")]
        defaultextension = ".xlsx"

        if extension == ".csv":
            filetypes = [("CSV 文件", "*.csv")]
            defaultextension = ".csv"

        path = filedialog.asksaveasfilename(
            title="选择导出位置",
            initialfile=os.path.basename(initial),
            initialdir=os.path.dirname(initial) or os.getcwd(),
            defaultextension=defaultextension,
            filetypes=filetypes,
        )

        if path:
            self.output_path.set(path)

    def choose_cache_import(self):
        path = filedialog.askopenfilename(
            title="选择要导入价格库的表格文件",
            filetypes=[
                ("表格文件", "*.xlsx *.xlsm *.csv"),
                ("Excel 文件", "*.xlsx *.xlsm"),
                ("CSV 文件", "*.csv"),
                ("所有文件", "*.*"),
            ],
        )

        if path:
            self.cache_import_path.set(path)

    def choose_combo_cache_import(self):
        path = filedialog.askopenfilename(
            title="选择要导入严选组合价格库的表格文件",
            filetypes=[
                ("表格文件", "*.xlsx *.xlsm *.csv"),
                ("Excel 文件", "*.xlsx *.xlsm"),
                ("CSV 文件", "*.csv"),
                ("所有文件", "*.*"),
            ],
        )

        if path:
            self.combo_cache_import_path.set(path)

    def choose_match_input(self):
        path = filedialog.askopenfilename(
            title="选择要匹配售价的表格文件",
            filetypes=[
                ("表格文件", "*.xlsx *.xlsm *.csv"),
                ("Excel 文件", "*.xlsx *.xlsm"),
                ("CSV 文件", "*.csv"),
                ("所有文件", "*.*"),
            ],
        )

        if not path:
            return

        self.match_input_path.set(path)
        self.match_output_path.set(default_price_match_output_path(path))

    def choose_match_output(self):
        initial = self.match_output_path.get() or default_price_match_output_path(self.match_input_path.get() or "匹配售价.xlsx")
        extension = os.path.splitext(initial)[1].lower()
        filetypes = [("Excel 文件", "*.xlsx")]
        defaultextension = ".xlsx"

        if extension == ".csv":
            filetypes = [("CSV 文件", "*.csv")]
            defaultextension = ".csv"

        path = filedialog.asksaveasfilename(
            title="选择匹配售价导出位置",
            initialfile=os.path.basename(initial),
            initialdir=os.path.dirname(initial) or os.getcwd(),
            defaultextension=defaultextension,
            filetypes=filetypes,
        )

        if path:
            self.match_output_path.set(path)

    def choose_rules_file(self):
        path = filedialog.askopenfilename(
            title="选择删除规则文件",
            initialdir=os.path.dirname(self.rules_path.get()) or app_base_dir(),
            filetypes=[
                ("JSON 规则文件", "*.json"),
                ("所有文件", "*.*"),
            ],
        )

        if path:
            self.rules_path.set(path)

    def create_default_rules_file(self):
        path = filedialog.asksaveasfilename(
            title="生成默认删除规则文件",
            initialdir=os.path.dirname(self.rules_path.get()) or app_base_dir(),
            initialfile="rules.json",
            defaultextension=".json",
            filetypes=[("JSON 规则文件", "*.json")],
        )

        if not path:
            return

        try:
            save_default_rules(path)
            self.rules_path.set(path)
            messagebox.showinfo("已生成", f"默认规则文件已生成：\n{path}")
        except Exception as error:
            messagebox.showerror("生成失败", str(error))

    def get_shipping_fees(self):
        fees = []

        for (label, _), variable in zip(SHIPPING_FEE_RULES, self.shipping_fee_vars):
            text = variable.get().strip()
            try:
                fees.append(float(text))
            except ValueError:
                raise ValueError(f"快递费规则“{label}”必须填写数字。")

        return fees

    def get_price_rules_from_vars(self, threshold_var, low_margin_var, high_margin_var, label):
        try:
            threshold = float(threshold_var.get().strip())
            low_margin = float(low_margin_var.get().strip())
            high_margin = float(high_margin_var.get().strip())
        except ValueError:
            raise ValueError(f"{label}里的分界值和毛利率都必须填写数字。")

        if not (0 <= low_margin < 1) or not (0 <= high_margin < 1):
            raise ValueError(f"{label}里的两个毛利率都必须大于等于 0，并且小于 1。")

        return threshold, low_margin, high_margin

    def get_price_rules(self):
        return self.get_price_rules_from_vars(
            self.price_threshold,
            self.low_price_margin,
            self.high_price_margin,
            "普通商品售价规则",
        )

    def get_combo_price_rules(self):
        return self.get_price_rules_from_vars(
            self.combo_price_threshold,
            self.combo_low_price_margin,
            self.combo_high_price_margin,
            "严选组合售价规则",
        )

    def start_processing(self):
        input_path = self.input_path.get().strip()
        output_path = self.output_path.get().strip()

        if not input_path or not os.path.exists(input_path):
            messagebox.showwarning("缺少文件", "请先选择要导入的表格文件。")
            return

        if not output_path:
            self.output_path.set(default_output_path(input_path))
            output_path = self.output_path.get()

        try:
            shipping_fees = self.get_shipping_fees()
            price_rules = self.get_price_rules()
            delete_rules = load_delete_rules(self.rules_path.get().strip())
            price_match_column = self.price_match_column.get().strip()
            if not price_match_column:
                raise ValueError("售价匹配列名不能为空。")
            price_cache = load_price_cache()
        except ValueError as error:
            messagebox.showwarning("规则错误", str(error))
            return
        except Exception as error:
            messagebox.showwarning("规则错误", f"规则或售价保存文件读取失败：{error}")
            return

        self._set_buttons_state("disabled")
        self.progress_bar.config(mode="indeterminate")
        self.progress_bar.start(10)
        self.progress.set(0)
        self.status.set("正在处理，请稍候...")

        thread = threading.Thread(
            target=self._process_in_thread,
            args=(input_path, output_path, shipping_fees, price_rules, delete_rules, price_match_column, price_cache),
            daemon=True,
        )
        thread.start()

    def start_price_cache_import(self):
        self._start_price_cache_import(
            self.cache_import_path.get().strip(),
            PRODUCT_CODE_COLUMN_NAME,
            WEIGHT_COLUMN_NAME,
            COST_PRICE_COLUMN_NAME,
            "价格库",
            "请先选择要导入价格库的表格文件。",
            self.get_price_rules,
        )

    def start_combo_price_cache_import(self):
        self._start_price_cache_import(
            self.combo_cache_import_path.get().strip(),
            COMBO_PRODUCT_CODE_COLUMN_NAME,
            COMBO_WEIGHT_COLUMN_NAME,
            COMBO_COST_PRICE_COLUMN_NAME,
            "严选组合价格库",
            "请先选择要导入严选组合价格库的表格文件。",
            self.get_combo_price_rules,
        )

    def _start_price_cache_import(
        self,
        input_path,
        product_code_column,
        weight_column,
        cost_price_column,
        progress_label,
        missing_file_message,
        price_rules_getter,
    ):
        if not input_path or not os.path.exists(input_path):
            messagebox.showwarning("缺少文件", missing_file_message)
            return

        try:
            shipping_fees = self.get_shipping_fees()
            price_rules = price_rules_getter()
            price_cache = load_price_cache()
        except ValueError as error:
            messagebox.showwarning("规则错误", str(error))
            return
        except Exception as error:
            messagebox.showwarning("读取失败", f"价格库读取失败：{error}")
            return

        self._set_buttons_state("disabled")
        self.progress_bar.config(mode="indeterminate")
        self.progress_bar.start(10)
        self.progress.set(0)
        self.status.set(f"正在计算并导入{progress_label}，请稍候...")

        thread = threading.Thread(
            target=self._import_price_cache_in_thread,
            args=(input_path, shipping_fees, price_rules, price_cache, product_code_column, weight_column, cost_price_column, progress_label),
            daemon=True,
        )
        thread.start()

    def start_price_match(self):
        input_path = self.match_input_path.get().strip()
        output_path = self.match_output_path.get().strip()

        if not input_path or not os.path.exists(input_path):
            messagebox.showwarning("缺少文件", "请先选择要匹配售价的新表格文件。")
            return

        if not output_path:
            self.match_output_path.set(default_price_match_output_path(input_path))
            output_path = self.match_output_path.get()

        try:
            price_cache = load_price_cache()
        except Exception as error:
            messagebox.showwarning("读取失败", f"价格库读取失败：{error}")
            return

        self._set_buttons_state("disabled")
        self.progress_bar.config(mode="indeterminate")
        self.progress_bar.start(10)
        self.progress.set(0)
        self.status.set("正在从价格库匹配售价，请稍候...")

        thread = threading.Thread(
            target=self._match_prices_in_thread,
            args=(input_path, output_path, price_cache),
            daemon=True,
        )
        thread.start()

    def _match_prices_in_thread(self, input_path, output_path, price_cache):
        try:
            extension = os.path.splitext(input_path)[1].lower()

            if extension == ".csv":
                summary = append_cached_prices_csv(input_path, output_path, self._set_progress, price_cache)
            elif extension in (".xlsx", ".xlsm"):
                summary = append_cached_prices_xlsx(input_path, output_path, self._set_progress, price_cache)
            else:
                raise ValueError("暂不支持该文件格式，请使用 .xlsx、.xlsm 或 .csv。")

            matched = sum(item[1] for item in summary)
            unmatched = sum(item[2] for item in summary)
            self.after(0, lambda: self._finish_match_success(output_path, matched, unmatched))
        except Exception as error:
            message = str(error)
            self.after(0, lambda: self._finish_error(message))

    def _import_price_cache_in_thread(
        self,
        input_path,
        shipping_fees,
        price_rules,
        price_cache,
        product_code_column,
        weight_column,
        cost_price_column,
        progress_label,
    ):
        try:
            extension = os.path.splitext(input_path)[1].lower()

            if extension == ".csv":
                summary = import_prices_to_cache_csv(
                    input_path,
                    self._set_progress,
                    shipping_fees,
                    price_rules,
                    price_cache,
                    product_code_column,
                    weight_column,
                    cost_price_column,
                    progress_label,
                )
            elif extension in (".xlsx", ".xlsm"):
                summary = import_prices_to_cache_xlsx(
                    input_path,
                    self._set_progress,
                    shipping_fees,
                    price_rules,
                    price_cache,
                    product_code_column,
                    weight_column,
                    cost_price_column,
                    progress_label,
                )
            else:
                raise ValueError("暂不支持该文件格式，请使用 .xlsx、.xlsm 或 .csv。")

            save_price_cache(price_cache)
            added = sum(item[1] for item in summary)
            updated = sum(item[2] for item in summary)
            skipped = sum(item[3] for item in summary)
            self.after(0, lambda: self._finish_cache_import_success(added, updated, skipped))
        except Exception as error:
            message = str(error)
            self.after(0, lambda: self._finish_error(message))

    def _process_in_thread(self, input_path, output_path, shipping_fees, price_rules, delete_rules, price_match_column, price_cache):
        try:
            extension = os.path.splitext(input_path)[1].lower()

            if extension == ".csv":
                summary = process_csv(
                    input_path,
                    output_path,
                    "after",
                    self._set_progress,
                    shipping_fees,
                    price_rules,
                    delete_rules,
                    price_match_column,
                    price_cache,
                )
            elif extension in (".xlsx", ".xlsm"):
                summary = process_xlsx(
                    input_path,
                    output_path,
                    "after",
                    self._set_progress,
                    shipping_fees,
                    price_rules,
                    delete_rules,
                    price_match_column,
                    price_cache,
                )
            else:
                raise ValueError("暂不支持该文件格式，请使用 .xlsx、.xlsm 或 .csv。")

            save_price_cache(price_cache)
            deleted = sum(item[2] for item in summary)
            kept = sum(item[1] for item in summary)
            matched = sum(item[3] for item in summary)
            saved = sum(item[4] for item in summary)
            self.after(0, lambda: self._finish_success(output_path, kept, deleted, matched, saved))
        except Exception as error:
            message = str(error)
            self.after(0, lambda: self._finish_error(message))

    def _set_status(self, text):
        self.after(0, lambda: self.status.set(text))

    def _set_progress(self, text, percent=None):
        def update():
            self.status.set(text)
            if percent is not None:
                self.progress_bar.stop()
                self.progress_bar.config(mode="determinate")
                self.progress.set(percent)
            elif str(self.progress_bar.cget("mode")) != "indeterminate":
                self.progress_bar.config(mode="indeterminate")
                self.progress_bar.start(10)

        self.after(0, update)

    def _set_buttons_state(self, state):
        self.cache_import_button.config(state=state)
        self.combo_cache_import_button.config(state=state)
        self.match_button.config(state=state)

    def _finish_success(self, output_path, kept, deleted, matched, saved):
        self._set_buttons_state("normal")
        self.progress_bar.stop()
        self.progress_bar.config(mode="determinate")
        self.progress.set(100)
        self.status.set(f"处理完成：保留 {kept} 行，删除 {deleted} 行，匹配保存售价 {matched} 行，新增保存售价 {saved} 行。")
        messagebox.showinfo(
            "处理完成",
            f"已导出：\n{output_path}\n\n匹配保存售价：{matched} 行\n新增保存售价：{saved} 行",
        )

    def _finish_match_success(self, output_path, matched, unmatched):
        self._set_buttons_state("normal")
        self.progress_bar.stop()
        self.progress_bar.config(mode="determinate")
        self.progress.set(100)
        self.status.set(f"售价匹配完成：匹配 {matched} 行，未匹配 {unmatched} 行。")
        messagebox.showinfo(
            "匹配完成",
            f"已导出：\n{output_path}\n\n匹配售价：{matched} 行\n未匹配：{unmatched} 行",
        )

    def _finish_cache_import_success(self, added, updated, skipped):
        self._set_buttons_state("normal")
        self.progress_bar.stop()
        self.progress_bar.config(mode="determinate")
        self.progress.set(100)
        self.status.set(f"价格库导入完成：新增 {added} 条，更新 {updated} 条，跳过 {skipped} 行。")
        messagebox.showinfo(
            "导入完成",
            f"价格库已更新。\n\n新增：{added} 条\n更新：{updated} 条\n跳过：{skipped} 行",
        )

    def _finish_error(self, message):
        self._set_buttons_state("normal")
        self.progress_bar.stop()
        self.progress_bar.config(mode="determinate")
        self.progress.set(0)
        self.status.set(f"处理失败：{message}")
        messagebox.showerror("处理失败", message)


if __name__ == "__main__":
    App().mainloop()
